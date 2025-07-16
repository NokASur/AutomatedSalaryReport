from salary_report_parser.worker import Worker, Job
from openpyxl import load_workbook
import logging
from logs.logging_module import logger, generate_handler
import os

log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")

os.makedirs(log_path, exist_ok=True)
logger.addHandler(generate_handler(os.path.join(log_path, "salary_report_parser_all.log"), logging.DEBUG))
logger.addHandler(generate_handler(os.path.join(log_path, "salary_report_parser_info.log"), logging.INFO))
logger.setLevel(logging.DEBUG)
logger.info(f"Logger enabled")


def safe_stoi_convertion(s: str | None) -> int | None:
    if s is None:
        return s
    try:
        return int(s)
    except (ValueError, TypeError) as e:
        logger.error(f"Fail during stoi convertion {e} from string: {s}")
        return None


# A slobby amoeba
def parse_excel_report(path: str) -> (dict[str, Worker], str):
    path = os.path.abspath(path)

    logger.info(f"Trying to parse: {path}")
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    last_sheet_name = sheet_names[-1]
    sheet = workbook[last_sheet_name]
    workers_dict = {}
    date = sheet.cell(row=3, column=3).value.strftime("%d-%m-%Y")

    for row in sheet.iter_rows(min_row=8, values_only=True):
        if row[0] is not None and row[1] is not None:
            # RE DO
            work_type = row[5]
            mark = row[6]
            tonns = row[7]
            runs = row[8]
            hectars = row[9]
            hours = row[10]
            salary_for_day = row[13]
            job = Job(
                work_type=work_type,
                mark=mark,
                tonns=tonns,
                runs=runs,
                hectars=hectars,
                hours=hours,
                salary_for_day=salary_for_day)

            worker = Worker(
                unique_id=row[1],
                name=row[2],
                machine_type=row[3],
                commentary=row[4],
                hours_worked_sum=round(safe_stoi_convertion(row[11]), 2) if safe_stoi_convertion(row[11]) else None,
                days_worked=row[12],
                salary_for_month=round(safe_stoi_convertion(row[14]), 2) if safe_stoi_convertion(row[14]) else None,
                repair_days_count=row[15],
                absence_reason=row[16],
                jobs=[job]
            )
            if workers_dict.get(worker.unique_id) is None:
                workers_dict[worker.unique_id] = worker
            else:
                workers_dict[worker.unique_id].add_jobs_from(worker)
        if row[1] is None:
            logger.info("Special data is skipped")
    logger.info(f"Parsed successfully")
    workbook.close()
    return workers_dict, date

# test

# workers, date = parse_excel_report("../unparsed_reports/report.xlsx")
# logger.info(f"Printing result messages:\n")
# for worker in workers:
#     logger.info(f"{worker.generate_message(date)}")
# logger.info(f"Finished printing result messages:\n")
#
# logger.info(f"Moving report to the ../parsed_reports/{date}report.xlsx folder:\n")
# shutil.move("../unparsed_reports/report.xlsx", f"../parsed_reports/{date}report.xlsx")
# logger.info(f"Moved successfully\n")
